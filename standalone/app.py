from __future__ import annotations

import io
import json
import os
import sys
import tempfile
from pathlib import Path

import time
import threading
from flask import Flask, jsonify, request, send_file, send_from_directory
from dotenv import load_dotenv

# PyInstaller 탐지용 명시적 임포트
try:
    import xlrd
    import openpyxl
except ImportError:
    pass

# PyInstaller 환경 대응을 위한 경로 처리
if hasattr(sys, "_MEIPASS"):
    _BASE = Path(sys._MEIPASS)
    _DIST = _BASE / "frontend" / "dist"
else:
    _BASE = Path(__file__).parent.parent
    _DIST = Path(__file__).parent / "frontend" / "dist"

sys.path.insert(0, str(_BASE))

from core.parser import parse_file
from core.template import download_template

load_dotenv()

app = Flask(__name__, static_folder=str(_DIST), static_url_path="")

# ── 하트비트 감시 로직 ───────────────────────────────────────────────────
LAST_HEARTBEAT = time.time()
HEARTBEAT_TIMEOUT = 15  # 15초 동안 신호 없으면 종료

@app.route("/api/heartbeat", methods=["POST"])
def heartbeat():
    global LAST_HEARTBEAT
    LAST_HEARTBEAT = time.time()
    return jsonify({"status": "ok"})

def watchdog():
    global LAST_HEARTBEAT
    while True:
        time.sleep(5)
        # 마지막 신호 이후 지정된 시간이 지나면 프로세스 강제 종료
        if time.time() - LAST_HEARTBEAT > HEARTBEAT_TIMEOUT:
            print("Heartbeat timeout. Shutting down...")
            os._exit(0)

# 백그라운드 스레드로 감시 시작
monitor_thread = threading.Thread(target=watchdog, daemon=True)
monitor_thread.start()


# ── 정적 파일 ─────────────────────────────────────────────────────────────

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_spa(path: str):
    target = _DIST / path
    if path and target.exists():
        return send_from_directory(str(_DIST), path)
    return send_from_directory(str(_DIST), "index.html")


# ── API ───────────────────────────────────────────────────────────────────

@app.route("/api/parse", methods=["POST"])
def api_parse():
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400

    f = request.files["file"]
    suffix = Path(f.filename).suffix.lower()
    if suffix not in (".xls", ".xlsx"):
        return jsonify({"error": "xls 또는 xlsx 파일만 허용됩니다."}), 400

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        tables = parse_file(tmp_path, infer_fk=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    return jsonify(tables)


@app.route("/api/template", methods=["GET"])
def api_template():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        download_template(tmp_path)
        return send_from_directory(
            os.path.dirname(tmp_path),
            os.path.basename(tmp_path),
            as_attachment=True,
            download_name="table_template.xlsx",
        )
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/excel", methods=["POST"])
def api_export_excel():
    import json as _json
    import io
    import xlsxwriter

    try:
        raw = request.get_data(as_text=True)
        body = _json.loads(raw) if raw else None
    except Exception as e:
        return jsonify({"error": f"JSON 파싱 실패: {e}"}), 400

    if not isinstance(body, dict):
        return jsonify({"error": f"payload가 dict가 아닙니다. type={type(body)}"}), 400

    if "nodes" not in body:
        return jsonify({"error": "노드 데이터가 없습니다."}), 400

    try:
        nodes = body.get("nodes", [])
        if not isinstance(nodes, list):
            nodes = [nodes] if nodes else []
            
        out = io.BytesIO()
        workbook = xlsxwriter.Workbook(out, {'in_memory': True})
        ws = workbook.add_worksheet("ERD")
        ws.hide_gridlines(2)
        
        node_map = {n.get("id"): n for n in nodes if isinstance(n, dict)}
        
        # Draw lines first so they appear "under" the boxes
        edges = body.get("edges", [])
        if not isinstance(edges, list):
            edges = []
            
        for edge in edges:
            if not isinstance(edge, dict):
                continue
            src_id = edge.get("source")
            tgt_id = edge.get("target")
            
            s_node = node_map.get(src_id)
            t_node = node_map.get(tgt_id)
            
            if s_node and t_node:
                # Calculate center points
                x1 = int(s_node.get("x", 0)) + int(s_node.get("width", 240)) // 2
                y1 = int(s_node.get("y", 0)) + int(s_node.get("height", 300)) // 2
                x2 = int(t_node.get("x", 0)) + int(t_node.get("width", 240)) // 2
                y2 = int(t_node.get("y", 0)) + int(t_node.get("height", 300)) // 2
                
                e_type = edge.get("edgeType", "inferred")
                color = '#94A3B8' # default inferred (greyish)
                if e_type == 'fk': color = '#86EFAC' # green
                elif e_type == 'custom': color = '#C084FC' # purple
                
                # Draw a simple line from center to center. 
                # Since boxes have fill color, the center-to-center line will naturally look like it's connecting the edges.
                # Draw orthogonal 'step' lines using skinny textboxes (since xlsxwriter doesn't support real lines)
                # Segment 1: Horizontal
                ws.insert_textbox('A1', '', {
                    'x_offset': min(x1, x2),
                    'y_offset': y1,
                    'width': max(2, abs(x2 - x1)),
                    'height': 2,
                    'fill': {'color': color},
                    'line': {'none': True},
                    'object_position': 1
                })
                # Segment 2: Vertical
                ws.insert_textbox('A1', '', {
                    'x_offset': x2,
                    'y_offset': min(y1, y2),
                    'width': 2,
                    'height': max(2, abs(y2 - y1)),
                    'fill': {'color': color},
                    'line': {'none': True},
                    'object_position': 1
                })

        # Calculate max boundaries to set a background color or just leave plain
        for idx, node in enumerate(nodes):
            if not isinstance(node, dict):
                continue
                
            x = int(node.get("x", 0))
            y = int(node.get("y", 0))
            w = int(node.get("width", 240))
            h = int(node.get("height", 300))
            
            data = node.get("data", {})
            if not isinstance(data, dict):
                data = {}
                
            name = data.get("table_name", node.get("id", ""))
            
            # Format the text content for the textbox
            content = f"[{name}]\n"
            cols = data.get("columns", [])
            if not isinstance(cols, list):
                cols = []
                
            for col in cols:
                if not isinstance(col, dict):
                    continue
                pk_str = "🔑" if col.get("pk") else ""
                fk_str = "FK" if col.get("fk_table") else ""
                flags = " ".join(filter(None, [pk_str, fk_str]))
                cname = col.get("name", "")
                ctype = col.get("type", "")
                content += f"{flags} {cname} : {ctype}\n"
            
            # Use insert_textbox to simulate table blocks
            options = {
                'x_offset': x,
                'y_offset': y,
                'width': w,
                'height': h,
                'fill': {'color': '#1E2330'},
                'font': {'color': '#93C5FD', 'name': 'Consolas', 'size': 9},
                'line': {'color': '#3B4A6B', 'width': 1.25},
                'align': {'vertical': 'top', 'horizontal': 'left'},
                'object_position': 1, # Move and size with cells
            }
            # Put them in cell A1 but offset via x/y so they float absolutely
            ws.insert_textbox('A1', content.strip(), options)

        workbook.close()
        out.seek(0)
        
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="erd_export.xlsx",
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Excel 생성 오류: {e}"}), 500


def _build_excel_image(img_bytes: bytes) -> bytes:
    import openpyxl
    from openpyxl.drawing.image import Image as XlImage

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERD"
    ws.sheet_view.showGridLines = False

    img_io = io.BytesIO(img_bytes)
    xl_img = XlImage(img_io)

    # pixelRatio:2 로 캡처된 이미지를 절반 크기로 조정 (화면 해상도에 맞춤)
    xl_img.width  = xl_img.width  // 2
    xl_img.height = xl_img.height // 2

    ws.add_image(xl_img, "A1")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _build_excel(tables: dict) -> bytes:
    import math
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERD"
    ws.sheet_view.showGridLines = False

    # ── 색상 ─────────────────────────────────────────────────────────────
    C_HEADER_BG  = "FF2D3A52"
    C_COLHDR_BG  = "FF1A2640"
    C_ROW_BG     = "FF111827"
    C_ROW_ALT    = "FF0F1520"
    C_PK_BG      = "FF1A1600"
    C_EMPTY_BG   = "FF0A0F18"
    C_HEADER_FG  = "FFBAE0F7"
    C_COLHDR_FG  = "FF64748B"
    C_PK_FG      = "FFFBBF24"
    C_TEXT       = "FFCBD5E1"
    C_TYPE       = "FF64748B"
    C_DESC       = "FF94A3B8"
    C_FK_FG      = "FF86EFAC"

    # ── 스타일 헬퍼 ──────────────────────────────────────────────────────
    def mk_fill(hex_color):
        return PatternFill(patternType="solid", fgColor=hex_color)

    def mk_font(color, bold=False, size=10, italic=False):
        return Font(color=color, bold=bold, size=size, italic=italic, name="Consolas")

    def mk_side(style, color="FF2D3A52"):
        return Side(style=style, color=color)

    THIN = mk_side("thin")
    MED  = mk_side("medium", "FF3B4A6B")

    def border(left=THIN, right=THIN, top=THIN, bottom=THIN):
        return Border(left=left, right=right, top=top, bottom=bottom)

    AL_C = Alignment(horizontal="center", vertical="center")
    AL_L = Alignment(horizontal="left",   vertical="center", indent=1)

    # ── 레이아웃 ─────────────────────────────────────────────────────────
    ids   = list(tables.keys())
    total = len(ids)
    if total == 0:
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return out.getvalue()

    TPROW = 4                  # 행당 최대 테이블 수
    COLS  = min(total, TPROW)
    TCOLS = 4                  # 테이블당 열 수 (PK, 필드ID, 타입, 설명)
    CGAP  = 2                  # 테이블 간 빈 열

    # 열 너비 설정
    for ci in range(COLS):
        base = 2 + ci * (TCOLS + CGAP)
        ws.column_dimensions[get_column_letter(base)    ].width = 4
        ws.column_dimensions[get_column_letter(base + 1)].width = 22
        ws.column_dimensions[get_column_letter(base + 2)].width = 14
        ws.column_dimensions[get_column_letter(base + 3)].width = 26

    # 각 행 블록의 시작 행 계산
    num_blocks = math.ceil(total / COLS)
    row_starts = []
    cur_row    = 2
    for rb in range(num_blocks):
        row_starts.append(cur_row)
        s = rb * COLS
        e = min(s + COLS, total)
        max_data = max(len(tables[ids[j]].get("columns", [])) for j in range(s, e))
        cur_row += 2 + max_data + 3   # header(1) + col-header(1) + data + gap(3)

    # ── 테이블 그리기 ────────────────────────────────────────────────────
    for i, tid in enumerate(ids):
        table   = tables[tid]
        columns = table.get("columns", [])
        rb      = i // COLS
        cb      = i % COLS
        sr      = row_starts[rb]
        sc      = 2 + cb * (TCOLS + CGAP)   # 시작 열 (1-indexed)

        s_blk = rb * COLS
        e_blk = min(s_blk + COLS, total)
        max_data = max(len(tables[ids[j]].get("columns", [])) for j in range(s_blk, e_blk))

        def cell_border(row_i, ci, total_rows):
            is_first = row_i == 0
            is_last  = row_i == total_rows - 1
            is_left  = ci == 0
            is_right = ci == TCOLS - 1
            return border(
                left   = MED  if is_left  else THIN,
                right  = MED  if is_right else THIN,
                top    = MED  if is_first else THIN,
                bottom = MED  if is_last  else THIN,
            )

        total_rows = 2 + max_data  # header + col-header + data rows

        # 헤더 행
        ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 3)
        hc = ws.cell(sr, sc)
        hc.value     = f"  {tid}  ·  {table.get('table_name', '')}"
        hc.fill      = mk_fill(C_HEADER_BG)
        hc.font      = mk_font(C_HEADER_FG, bold=True, size=11)
        hc.alignment = AL_L
        for ci in range(TCOLS):
            c = ws.cell(sr, sc + ci)
            c.fill   = mk_fill(C_HEADER_BG)
            c.border = cell_border(0, ci, total_rows)

        # 컬럼 헤더 행
        col_hdrs = ["PK", "Field ID", "Data Type", "Description"]
        for ci, h in enumerate(col_hdrs):
            c = ws.cell(sr + 1, sc + ci)
            c.value     = h
            c.fill      = mk_fill(C_COLHDR_BG)
            c.font      = mk_font(C_COLHDR_FG, bold=True, size=9)
            c.alignment = AL_C if ci == 0 else AL_L
            c.border    = cell_border(1, ci, total_rows)

        # 데이터 행
        for di, col in enumerate(columns):
            row    = sr + 2 + di
            is_pk  = col.get("pk", False)
            is_fk  = bool(col.get("fk_table")) and not is_pk
            row_bg = C_PK_BG if is_pk else (C_ROW_ALT if di % 2 == 0 else C_ROW_BG)
            ri     = 2 + di   # row index within table block

            vals  = [
                "🔑" if is_pk else ("FK" if is_fk else ""),
                col.get("name", ""),
                col.get("type", ""),
                col.get("description", ""),
            ]
            fonts = [
                mk_font(C_PK_FG, bold=True) if is_pk else mk_font(C_FK_FG, size=9) if is_fk else mk_font(C_TEXT, size=9),
                mk_font(C_PK_FG, bold=True) if is_pk else mk_font(C_TEXT),
                mk_font(C_TYPE, size=9, italic=True),
                mk_font(C_DESC, size=9),
            ]
            for ci in range(TCOLS):
                c = ws.cell(row, sc + ci)
                c.value     = vals[ci]
                c.fill      = mk_fill(row_bg)
                c.font      = fonts[ci]
                c.alignment = AL_C if ci == 0 else AL_L
                c.border    = cell_border(ri, ci, total_rows)

        # 데이터가 이 테이블보다 많은 경우 빈 행으로 채움
        for di in range(len(columns), max_data):
            row = sr + 2 + di
            ri  = 2 + di
            for ci in range(TCOLS):
                c = ws.cell(row, sc + ci)
                c.fill   = mk_fill(C_EMPTY_BG)
                c.border = cell_border(ri, ci, total_rows)

    # 행 높이 설정
    ws.row_dimensions[1].height = 6
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── 진입점 ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import webbrowser
    import argparse
    from threading import Timer

    parser = argparse.ArgumentParser(description="ERD Tool")
    parser.add_argument("-p", "--port", type=int, default=5000, help="Port to run the server on (default: 5000)")
    args = parser.parse_args()

    port = args.port

    def open_browser():
        webbrowser.open_new(f"http://127.0.0.1:{port}")

    # EXE 실행 시 또는 디버그 모드가 아닐 때 브라우저 자동 실행
    if not os.environ.get("WERKZEUG_RUN_MAIN"):
        Timer(1.5, open_browser).start()
    
    app.run(host="127.0.0.1", port=port, debug=False)
