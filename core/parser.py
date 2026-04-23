from __future__ import annotations

from pathlib import Path
from typing import Any, Protocol

SKIP_SHEETS = {"T_00", "목록"}

COL_FIELD_ID = 1
COL_DESCRIPTION = 2
COL_PK = 3
COL_DATA_TYPE = 4
COL_NULL = 5
COL_DEFAULT = 6
COL_FK_TABLE = 7
COL_FK_COLUMN = 8

ROW_TABLE_ID = 2
ROW_TABLE_SPACE = 3
ROW_TABLE_NAME = 4
ROW_COLUMNS_START = 7
ROW_INDEX_KEY = 38


# ── 공통 Sheet 어댑터 ─────────────────────────────────────────────────────

class SheetAdapter(Protocol):
    @property
    def nrows(self) -> int: ...
    def cell_value(self, row: int, col: int) -> Any: ...
    def sheet_names(self) -> list[str]: ...  # workbook-level, not used here


class XlrdSheetAdapter:
    def __init__(self, sheet: Any):
        self._sheet = sheet

    @property
    def nrows(self) -> int:
        return self._sheet.nrows

    def cell_value(self, row: int, col: int) -> Any:
        try:
            return self._sheet.cell_value(row, col)
        except IndexError:
            return None


class OpenpyxlSheetAdapter:
    def __init__(self, sheet: Any):
        self._sheet = sheet

    @property
    def nrows(self) -> int:
        return self._sheet.max_row

    def cell_value(self, row: int, col: int) -> Any:
        # openpyxl은 1-indexed
        try:
            return self._sheet.cell(row + 1, col + 1).value
        except Exception:
            return None


# ── 공통 파싱 함수 ────────────────────────────────────────────────────────

def _is_numeric(val: Any) -> bool:
    try:
        float(val)
        return True
    except (TypeError, ValueError):
        return False


def _is_pk(val: Any) -> bool:
    """D열 PK 컬럼 값이 Y/y 또는 숫자 1이면 PK로 인식."""
    if val in ("", None):
        return False
    s = str(val).strip()
    if s.upper() in ("Y", "YES"):
        return True
    try:
        return float(s) != 0
    except (TypeError, ValueError):
        return False


def _cell_str(sheet: Any, row: int, col: int) -> str:
    val = sheet.cell_value(row, col)
    return str(val).strip() if val not in ("", None) else ""


def _parse_columns(sheet: Any) -> list[dict]:
    columns = []
    for i in range(ROW_COLUMNS_START, sheet.nrows):
        no_val = sheet.cell_value(i, 0)
        if not _is_numeric(no_val):
            break
        field_id = _cell_str(sheet, i, COL_FIELD_ID)
        if not field_id:
            continue
        pk_val = sheet.cell_value(i, COL_PK)
        fk_table = _cell_str(sheet, i, COL_FK_TABLE) or None
        fk_column = _cell_str(sheet, i, COL_FK_COLUMN) or None
        columns.append({
            "name": field_id,
            "description": _cell_str(sheet, i, COL_DESCRIPTION),
            "type": _cell_str(sheet, i, COL_DATA_TYPE),
            "pk": _is_pk(pk_val),
            "null": _cell_str(sheet, i, COL_NULL),
            "default": _cell_str(sheet, i, COL_DEFAULT),
            "fk_table": fk_table,
            "fk_column": fk_column,
        })
    return columns


def _parse_index_key(sheet: Any) -> list[dict]:
    """INDEX/KEY 섹션 파싱 — 포맷 미확정이므로 graceful 처리."""
    relations = []
    if sheet.nrows <= ROW_INDEX_KEY:
        return relations
    for i in range(ROW_INDEX_KEY + 1, sheet.nrows):
        kind = _cell_str(sheet, i, 0).upper()
        col_name = _cell_str(sheet, i, 1)
        ref_table = _cell_str(sheet, i, 2)
        ref_col = _cell_str(sheet, i, 3)
        if not kind or not col_name:
            continue
        if kind == "FK" and ref_table:
            relations.append({
                "type": "FK",
                "column": col_name,
                "ref_table": ref_table,
                "ref_column": ref_col,
            })
    return relations


def _build_inline_fk_relations(columns: list[dict]) -> list[dict]:
    """H열(FK Table) / I열(FK Column) 명시 FK → relations 목록"""
    relations = []
    for col in columns:
        fk_table = col.get("fk_table")
        if not fk_table:
            continue
        fk_column = col.get("fk_column") or col["name"]
        relations.append({
            "type": "FK",
            "column": col["name"],
            "ref_table": fk_table,
            "ref_column": fk_column,
        })
    return relations


def _infer_fk_from_columns(columns: list[dict]) -> list[dict]:
    """컬럼명 패턴(_CD, _ID 등)으로 FK 보조 추론."""
    FK_SUFFIXES = ("_CD", "_ID", "_NO", "_SEQ")
    relations = []
    for col in columns:
        if col["pk"]:
            continue
        name_upper = col["name"].upper()
        for suffix in FK_SUFFIXES:
            if name_upper.endswith(suffix):
                relations.append({
                    "type": "FK_INFERRED",
                    "column": col["name"],
                    "ref_table": None,
                    "ref_column": col["name"],
                })
                break
    return relations


# ── 워크북 로더 ───────────────────────────────────────────────────────────

def _load_workbook(file_path: str) -> tuple[list[str], dict[str, Any], Any]:
    """확장자에 따라 xlrd / openpyxl로 로드. (sheet_names, name→adapter, book) 반환."""
    ext = Path(file_path).suffix.lower()
    if ext == ".xls":
        import xlrd
        book = xlrd.open_workbook(file_path)
        adapters = {
            name: XlrdSheetAdapter(book.sheet_by_name(name))
            for name in book.sheet_names()
        }
        return book.sheet_names(), adapters, book
    elif ext == ".xlsx":
        import openpyxl
        book = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        adapters = {
            name: OpenpyxlSheetAdapter(book[name])
            for name in book.sheetnames
        }
        return book.sheetnames, adapters, book
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {ext} (xls / xlsx만 허용)")


# ── 공개 API ──────────────────────────────────────────────────────────────

def parse_file(file_path: str, infer_fk: bool = True) -> dict[str, dict]:
    """XLS / XLSX 파일 파싱 → 테이블 객체 dict 반환."""
    sheet_names, adapters, book = _load_workbook(file_path)

    tables: dict[str, dict] = {}

    for name in sheet_names:
        if name in SKIP_SHEETS:
            continue
        sheet = adapters[name]

        table_id = _cell_str(sheet, ROW_TABLE_ID, 2)
        table_space = _cell_str(sheet, ROW_TABLE_SPACE, 2)
        table_name = _cell_str(sheet, ROW_TABLE_NAME, 2)

        if not table_id:
            continue

        columns = _parse_columns(sheet)
        inline_relations = _build_inline_fk_relations(columns)
        index_key_relations = _parse_index_key(sheet)
        covered = {r["column"] for r in inline_relations}
        extra = [r for r in index_key_relations if r["column"] not in covered]

        tables[table_id] = {
            "table_id": table_id,
            "table_name": table_name,
            "table_space": table_space,
            "columns": columns,
            "relations": inline_relations + extra,
        }

    if hasattr(book, "close"):
        book.close()

    known_ids = set(tables.keys())

    # 명시 FK의 ref_table이 실제 존재하는 테이블 ID인지 검증 (비고 열 오독 방지)
    for tbl in tables.values():
        tbl["relations"] = [
            r for r in tbl["relations"]
            if r["type"] != "FK" or r.get("ref_table") in known_ids
        ]

    if infer_fk:
        # 명시 FK가 없는 테이블에만 컬럼명 패턴 추론 적용
        for tbl in tables.values():
            if not tbl["relations"]:
                tbl["relations"] = _infer_fk_from_columns(tbl["columns"])
        _resolve_inferred_fk(tables)

    return tables


def _resolve_inferred_fk(tables: dict[str, dict]) -> None:
    """FK_INFERRED 관계의 ref_table을 전체 테이블의 PK 컬럼과 매칭해 채운다."""
    # 컬럼명(대문자) → PK를 가진 테이블 ID 매핑
    pk_col_to_table: dict[str, str] = {}
    for tid, tbl in tables.items():
        for col in tbl["columns"]:
            if col["pk"]:
                pk_col_to_table[col["name"].upper()] = tid

    for tid, tbl in tables.items():
        resolved = []
        for rel in tbl["relations"]:
            if rel["type"] == "FK_INFERRED" and rel["ref_table"] is None:
                ref_table = pk_col_to_table.get(rel["column"].upper())
                if ref_table and ref_table != tid:
                    resolved.append({**rel, "ref_table": ref_table, "ref_column": rel["column"]})
                # ref_table 못 찾으면 드롭 (그릴 수 없음)
            else:
                resolved.append(rel)
        tbl["relations"] = resolved


# 하위 호환 별칭
parse_xls = parse_file
